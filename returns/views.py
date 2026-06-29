import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    ReturnRequest,
    ReturnRequestHistory,
    ReturnRequestAttachment,
    ReturnRequestSerial,
)
from .forms import (
    ReturnRequestForm,
    ReturnClientResponseForm,
    ClienteUserCreateForm,
)


# ==========================================================
# ROLES / HELPERS GENERALES
# ==========================================================

def es_cliente(user):
    """
    Un usuario staff/superuser nunca debe comportarse como Cliente.
    """
    if not user.is_authenticated:
        return False

    if user.is_staff or user.is_superuser:
        return False

    return user.groups.filter(name__iexact='Cliente').exists()


def nombre_responsable_usuario(user):
    """
    Devuelve nombre completo si existe; si no, username.
    """
    nombre = user.get_full_name()
    return nombre if nombre else user.username


def valor_legible(valor):
    if valor is None or valor == '':
        return 'vacío'

    return str(valor)


def registrar_historial(devolucion, titulo, descripcion, usuario=None, tipo=None):
    ReturnRequestHistory.objects.create(
        return_request=devolucion,
        titulo=titulo,
        descripcion=descripcion,
        creado_por=usuario if usuario and usuario.is_authenticated else None,
        tipo=tipo
    )


def registrar_cambio_si_corresponde(devolucion, campo_nombre, etiqueta, valor_anterior, valor_nuevo, usuario):
    anterior = valor_legible(valor_anterior)
    nuevo = valor_legible(valor_nuevo)

    if anterior != nuevo:
        registrar_historial(
            devolucion=devolucion,
            titulo=f'{etiqueta} actualizado',
            descripcion=f'{etiqueta}: {anterior} → {nuevo}',
            usuario=usuario,
            tipo='CAMBIO_CLIENTE'
        )


def redirect_after_action(request, devolucion):
    next_url = request.POST.get('next') or request.GET.get('next')

    if next_url:
        return redirect(next_url)

    return redirect('returns:return_detail', pk=devolucion.pk)


# ==========================================================
# SERIES ILIMITADAS
# ==========================================================

def normalizar_serie(valor):
    """
    La serie se guarda SIEMPRE como texto.
    No se convierte a int para no perder ceros iniciales.
    Ej: 00299695106010016175 se mantiene igual.
    """
    if valor is None:
        return ''

    return str(valor).strip()


def obtener_series_desde_request(request):
    """
    Lee series desde:
    - series_json: lista JSON enviada por el JS del modal.
    - serie: campo antiguo/respaldo.

    No limita por cantidad.
    Elimina duplicados en el mismo envío, conservando el orden.
    """
    series = []

    raw = request.POST.get('series_json', '')

    if raw:
        try:
            data = json.loads(raw)

            if isinstance(data, list):
                for item in data:
                    serie = normalizar_serie(item)

                    if serie:
                        series.append(serie)
        except json.JSONDecodeError:
            pass

    serie_legacy = normalizar_serie(request.POST.get('serie'))

    if serie_legacy:
        series.append(serie_legacy)

    resultado = []
    vistos = set()

    for serie in series:
        key = serie.upper()

        if key in vistos:
            continue

        vistos.add(key)
        resultado.append(serie)

    return resultado


def guardar_series_devolucion(devolucion, request):
    """
    Guarda todas las series capturadas en ReturnRequestSerial.
    No hay límite.
    No pisa las series existentes.
    Mantiene devolucion.serie como respaldo con la primera serie.
    """
    series = obtener_series_desde_request(request)

    if not series:
        return 0

    existentes = set(
        str(serie).upper()
        for serie in devolucion.series.values_list('serie', flat=True)
    )

    total_creadas = 0

    for serie in series:
        key = serie.upper()

        if key in existentes:
            continue

        ReturnRequestSerial.objects.create(
            return_request=devolucion,
            serie=serie,
            creado_por=request.user
        )

        existentes.add(key)
        total_creadas += 1

    if not devolucion.serie and series:
        devolucion.serie = series[0]
        devolucion.save(update_fields=['serie', 'actualizado_en'])

    if total_creadas:
        registrar_historial(
            devolucion=devolucion,
            titulo='Series registradas',
            descripcion=f'Se registraron {total_creadas} serie(s): {", ".join(series)}',
            usuario=request.user,
            tipo='SERIES'
        )

    return total_creadas


def resumen_series_devolucion(devolucion):
    """
    Texto corto para tabla, detalle y API.
    """
    try:
        series = list(devolucion.series.values_list('serie', flat=True))
    except Exception:
        series = []

    if series:
        if len(series) <= 3:
            return ', '.join(series)

        return f'{len(series)} series: {", ".join(series[:3])}...'

    return devolucion.serie or '-'


def texto_series_exportacion(devolucion):
    """
    Texto completo para Excel.
    Una serie por línea.
    """
    try:
        series = list(devolucion.series.values_list('serie', flat=True))
    except Exception:
        series = []

    if series:
        return '\n'.join(series)

    return devolucion.serie or ''


# ==========================================================
# ADJUNTOS / GALERÍA
# ==========================================================

def guardar_adjuntos_internos(devolucion, request):
    documentos = request.FILES.getlist('documentos_adjuntos')
    fotos = request.FILES.getlist('fotos_evidencia')

    total_documentos = 0
    total_fotos = 0

    for archivo in documentos:
        ReturnRequestAttachment.objects.create(
            return_request=devolucion,
            archivo=archivo,
            tipo='DOCUMENTO_INTERNO',
            subido_por=request.user
        )
        total_documentos += 1

    for archivo in fotos:
        ReturnRequestAttachment.objects.create(
            return_request=devolucion,
            archivo=archivo,
            tipo='FOTO_INTERNA',
            subido_por=request.user
        )
        total_fotos += 1

    if total_documentos or total_fotos:
        registrar_historial(
            devolucion=devolucion,
            titulo='Adjuntos internos cargados',
            descripcion=f'Se cargaron {total_documentos} documento(s) y {total_fotos} foto(s) internas.',
            usuario=request.user,
            tipo='ADJUNTOS_INTERNOS'
        )


def archivo_data(archivo, titulo, tipo):
    """
    Convierte FileField/ImageField a estructura segura para el frontend.
    """
    if not archivo:
        return None

    try:
        url = archivo.url
    except Exception:
        return None

    nombre = archivo.name.split('/')[-1] if getattr(archivo, 'name', '') else titulo

    return {
        'titulo': titulo,
        'nombre': nombre,
        'url': url,
        'tipo': tipo,
    }


def obtener_galeria_devolucion(devolucion):
    """
    Devuelve documentos/fotos asociados a la devolución:
    - Campos antiguos.
    - Modelo ReturnRequestAttachment.
    """
    galeria = []

    archivos_principales = [
        ('documento_adjunto', 'Documento interno', 'document'),
        ('foto_evidencia', 'Foto interna', 'image'),
        ('documento_respuesta_cliente', 'Documento cliente', 'document'),
        ('foto_respuesta_cliente', 'Foto cliente', 'image'),
    ]

    for campo, titulo, tipo in archivos_principales:
        archivo = getattr(devolucion, campo, None)
        data = archivo_data(archivo, titulo, tipo)

        if data:
            galeria.append(data)

    if hasattr(devolucion, 'adjuntos'):
        for adjunto in devolucion.adjuntos.all().order_by('creado_en', 'id'):
            archivo = getattr(adjunto, 'archivo', None)
            tipo_adjunto = getattr(adjunto, 'tipo', '') or ''

            if not archivo:
                continue

            tipo_upper = tipo_adjunto.upper()
            es_foto = 'FOTO' in tipo_upper or 'IMAGE' in tipo_upper

            if 'CLIENTE' in tipo_upper:
                titulo = 'Foto cliente' if es_foto else 'Documento cliente'
            else:
                titulo = 'Foto interna' if es_foto else 'Documento interno'

            data = archivo_data(
                archivo,
                titulo,
                'image' if es_foto else 'document'
            )

            if data:
                galeria.append(data)

    return galeria


# ==========================================================
# FILTROS
# ==========================================================

def obtener_devoluciones_filtradas(request):
    query = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    base_queryset = (
        ReturnRequest.objects
        .all()
        .prefetch_related('series', 'adjuntos')
        .order_by('-fecha_recepcion', '-id')
    )

    devoluciones = base_queryset

    if fecha_desde:
        devoluciones = devoluciones.filter(fecha_recepcion__gte=fecha_desde)

    if fecha_hasta:
        devoluciones = devoluciones.filter(fecha_recepcion__lte=fecha_hasta)

    if query:
        devoluciones = devoluciones.filter(
            Q(numero_documento__icontains=query) |
            Q(documento_ingreso__icontains=query) |
            Q(cliente__icontains=query) |
            Q(sku__icontains=query) |
            Q(codigo_numerico__icontains=query) |
            Q(serie__icontains=query) |
            Q(series__serie__icontains=query) |
            Q(ubicacion__icontains=query) |
            Q(numero_dcto_estado__icontains=query) |
            Q(responsable__icontains=query)
        ).distinct()

    if estado:
        devoluciones = devoluciones.filter(estado=estado)

    return base_queryset, devoluciones, query, estado, fecha_desde, fecha_hasta


# ==========================================================
# VISTAS PRINCIPALES
# ==========================================================

@login_required
def return_list(request):
    base_queryset, devoluciones, query, estado, fecha_desde, fecha_hasta = obtener_devoluciones_filtradas(request)
    cliente = es_cliente(request.user)

    return render(request, 'returns/return_list.html', {
        'devoluciones': devoluciones,
        'query': query,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'estado_choices': ReturnRequest.ESTADO_CHOICES,

        'form': ReturnRequestForm(user=request.user),
        'user_form': ClienteUserCreateForm(),

        'open_user_modal': False,
        'es_cliente': cliente,

        'total_devoluciones': base_queryset.count(),
        'total_recibidas': devoluciones.filter(estado='RECIBIDO').count(),
        'total_revision': devoluciones.filter(estado='EN_REVISION').count(),
        'total_observadas': devoluciones.filter(estado='OBSERVADO').count(),
        'total_aprobadas': devoluciones.filter(estado='APROBADO').count(),
        'total_rechazadas': devoluciones.filter(estado='RECHAZADO').count(),
        'total_cerradas': devoluciones.filter(estado='CERRADO').count(),
        'total_filtradas': devoluciones.count(),
    })


@login_required
def return_create(request):
    if es_cliente(request.user):
        raise PermissionDenied

    if request.method == 'POST':
        form = ReturnRequestForm(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            cantidad = form.cleaned_data.get('cantidad') or 0
            cantidad = int(cantidad)

            # Lee las series originales enviadas desde el modal
            series_raw = []
            raw_json = request.POST.get('series_json', '')

            if raw_json:
                try:
                    data = json.loads(raw_json)

                    if isinstance(data, list):
                        for item in data:
                            serie = normalizar_serie(item)
                            if serie:
                                series_raw.append(serie)

                except json.JSONDecodeError:
                    messages.error(request, 'Error leyendo las series capturadas. Intente nuevamente.')
                    return redirect('returns:return_list')

            # También toma el campo legacy "serie" si viene lleno
            serie_legacy = normalizar_serie(request.POST.get('serie'))

            if serie_legacy:
                series_raw.append(serie_legacy)

            # Validar duplicados antes de guardar
            vistos = set()
            duplicadas = []

            for serie in series_raw:
                key = serie.upper()

                if key in vistos:
                    duplicadas.append(serie)
                else:
                    vistos.add(key)

            if duplicadas:
                messages.error(
                    request,
                    f'Serie duplicada detectada: {", ".join(duplicadas)}. No puedes registrar series repetidas.'
                )
                return redirect('returns:return_list')

            # Series normalizadas y únicas
            series = obtener_series_desde_request(request)
            total_series = len(series)

            # Validar que la cantidad coincida exactamente con las series
            if total_series < cantidad:
                faltantes = cantidad - total_series

                messages.error(
                    request,
                    f'Faltan {faltantes} serie(s) por capturar. '
                    f'Cantidad indicada: {cantidad}. Series capturadas: {total_series}.'
                )
                return redirect('returns:return_list')

            if total_series > cantidad:
                sobrantes = total_series - cantidad

                messages.error(
                    request,
                    f'Tienes {sobrantes} serie(s) de más. '
                    f'Cantidad indicada: {cantidad}. Series capturadas: {total_series}.'
                )
                return redirect('returns:return_list')

            devolucion = form.save(commit=False)
            devolucion.estado = 'RECIBIDO'
            devolucion.creado_por = request.user
            devolucion.responsable = nombre_responsable_usuario(request.user)
            devolucion.save()

            guardar_adjuntos_internos(devolucion, request)
            total_series_guardadas = guardar_series_devolucion(devolucion, request)

            messages.success(
                request,
                f'Devolución registrada correctamente con {total_series_guardadas} serie(s).'
            )

            return redirect('returns:return_list')

        messages.error(request, 'No se pudo registrar la devolución. Revisa los campos obligatorios.')

    else:
        form = ReturnRequestForm(user=request.user)

    return render(request, 'returns/return_form.html', {
        'form': form,
        'title': 'Registrar devolución',
        'es_cliente': es_cliente(request.user),
    })


@login_required
def return_detail(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)
    cliente = es_cliente(request.user)

    if cliente and devolucion.estado == 'RECIBIDO':
        ahora = timezone.now()

        devolucion.estado = 'EN_REVISION'

        if not devolucion.fecha_en_revision:
            devolucion.fecha_en_revision = ahora

        devolucion.save(update_fields=[
            'estado',
            'fecha_en_revision',
            'actualizado_en'
        ])

        registrar_historial(
            devolucion=devolucion,
            titulo='Estado actualizado',
            descripcion='Estado: Recibido → En revisión',
            usuario=request.user,
            tipo='EN_REVISION'
        )

    return render(request, 'returns/return_detail.html', {
        'devolucion': devolucion,
        'es_cliente': cliente,
        'response_form': ReturnClientResponseForm(
            instance=devolucion,
            initial={'respuesta_cliente': ''}
        ),
        'historial_extra': devolucion.historial.all(),
        'documentos_internos': devolucion.adjuntos.filter(tipo='DOCUMENTO_INTERNO'),
        'fotos_internas': devolucion.adjuntos.filter(tipo='FOTO_INTERNA'),
    })


@login_required
def return_update(request, pk):
    """
    Cliente no puede editar devolución completa.
    Admin puede editar mientras no esté en estado final.
    También puede agregar series nuevas desde el mismo request.
    """
    if es_cliente(request.user):
        raise PermissionDenied

    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if devolucion.estado in ['APROBADO', 'RECHAZADO', 'CERRADO']:
        messages.warning(request, 'Esta devolución ya tiene un estado final y no puede editarse.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        form = ReturnRequestForm(
            request.POST,
            request.FILES,
            instance=devolucion,
            user=request.user
        )

        if form.is_valid():
            devolucion = form.save()
            guardar_series_devolucion(devolucion, request)
            guardar_adjuntos_internos(devolucion, request)

            messages.success(request, 'Devolución actualizada correctamente.')
            return redirect_after_action(request, devolucion)
    else:
        form = ReturnRequestForm(instance=devolucion, user=request.user)

    return render(request, 'returns/return_form.html', {
        'form': form,
        'title': 'Editar devolución',
        'devolucion': devolucion,
        'es_cliente': es_cliente(request.user),
    })


# ==========================================================
# ACCIONES CLIENTE / ADMIN
# ==========================================================

@login_required
def return_respond(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if not es_cliente(request.user):
        raise PermissionDenied

    if devolucion.estado in ['APROBADO', 'RECHAZADO', 'CERRADO']:
        messages.warning(request, 'Esta devolución ya tiene un estado final.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        documento_ingreso_anterior = devolucion.documento_ingreso
        estado_inventario_anterior = devolucion.numero_dcto_estado

        documento_cliente_anterior = (
            devolucion.documento_respuesta_cliente.name
            if devolucion.documento_respuesta_cliente else ''
        )

        foto_cliente_anterior = (
            devolucion.foto_respuesta_cliente.name
            if devolucion.foto_respuesta_cliente else ''
        )

        form = ReturnClientResponseForm(
            request.POST,
            request.FILES,
            instance=devolucion
        )

        if form.is_valid():
            respuesta = form.save(commit=False)
            ahora = timezone.now()

            observacion_nueva = form.cleaned_data.get('respuesta_cliente') or ''
            observacion_nueva = observacion_nueva.strip()

            respuesta.estado = 'APROBADO'
            respuesta.respondido_por = request.user
            respuesta.fecha_respuesta_cliente = ahora
            respuesta.fecha_aprobacion_cliente = ahora
            respuesta.respuesta_cliente = observacion_nueva

            if not respuesta.fecha_en_revision:
                respuesta.fecha_en_revision = ahora

            if observacion_nueva:
                respuesta.fecha_observacion_cliente = ahora

            if (
                request.FILES.get('documento_respuesta_cliente') or
                request.FILES.get('foto_respuesta_cliente')
            ):
                respuesta.fecha_adjuntos_cliente = ahora

            respuesta.save()

            if observacion_nueva:
                descripcion_respuesta = f'El cliente envió respuesta. Observación: {observacion_nueva}'
            else:
                descripcion_respuesta = 'El cliente envió respuesta sin observaciones adicionales.'

            registrar_historial(
                devolucion=respuesta,
                titulo='Respuesta enviada por cliente',
                descripcion=descripcion_respuesta,
                usuario=request.user,
                tipo='RESPUESTA_CLIENTE'
            )

            registrar_cambio_si_corresponde(
                respuesta,
                'documento_ingreso',
                'Documento de ingreso / NC cliente',
                documento_ingreso_anterior,
                respuesta.documento_ingreso,
                request.user
            )

            registrar_cambio_si_corresponde(
                respuesta,
                'numero_dcto_estado',
                'Estado de Inventario',
                estado_inventario_anterior,
                respuesta.numero_dcto_estado,
                request.user
            )

            documento_cliente_nuevo = (
                respuesta.documento_respuesta_cliente.name
                if respuesta.documento_respuesta_cliente else ''
            )

            foto_cliente_nueva = (
                respuesta.foto_respuesta_cliente.name
                if respuesta.foto_respuesta_cliente else ''
            )

            registrar_cambio_si_corresponde(
                respuesta,
                'documento_respuesta_cliente',
                'Documento adjunto cliente',
                documento_cliente_anterior,
                documento_cliente_nuevo,
                request.user
            )

            registrar_cambio_si_corresponde(
                respuesta,
                'foto_respuesta_cliente',
                'Foto evidencia cliente',
                foto_cliente_anterior,
                foto_cliente_nueva,
                request.user
            )

            registrar_historial(
                devolucion=respuesta,
                titulo='Devolución aprobada',
                descripcion='El cliente aprobó la devolución.',
                usuario=request.user,
                tipo='APROBADO'
            )

            messages.success(request, 'Devolución aprobada correctamente.')
            return redirect_after_action(request, devolucion)

    messages.error(request, 'No se pudo registrar la respuesta. Revise los campos.')
    return redirect_after_action(request, devolucion)


@login_required
def return_reject(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if not es_cliente(request.user):
        raise PermissionDenied

    if devolucion.estado in ['APROBADO', 'RECHAZADO', 'CERRADO']:
        messages.warning(request, 'Esta devolución ya tiene un estado final.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        motivo_rechazo = request.POST.get('motivo_rechazo', '').strip()
        ahora = timezone.now()

        observacion_anterior = devolucion.respuesta_cliente

        devolucion.estado = 'RECHAZADO'
        devolucion.respondido_por = request.user
        devolucion.fecha_respuesta_cliente = ahora
        devolucion.fecha_rechazo_cliente = ahora

        if not devolucion.fecha_en_revision:
            devolucion.fecha_en_revision = ahora

        if motivo_rechazo:
            devolucion.respuesta_cliente = motivo_rechazo
            devolucion.fecha_observacion_cliente = ahora

        devolucion.save()

        registrar_cambio_si_corresponde(
            devolucion,
            'respuesta_cliente',
            'Motivo de rechazo',
            observacion_anterior,
            devolucion.respuesta_cliente,
            request.user
        )

        registrar_historial(
            devolucion=devolucion,
            titulo='Devolución rechazada',
            descripcion='El cliente rechazó la devolución.',
            usuario=request.user,
            tipo='RECHAZADO'
        )

        messages.success(request, 'Devolución rechazada correctamente.')
        return redirect_after_action(request, devolucion)

    return redirect_after_action(request, devolucion)


@login_required
def return_observe(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if not es_cliente(request.user):
        raise PermissionDenied

    if devolucion.estado in ['APROBADO', 'RECHAZADO', 'CERRADO']:
        messages.warning(request, 'Esta devolución ya tiene un estado final.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        observacion = request.POST.get('observacion_cliente', '').strip()

        if not observacion:
            messages.error(request, 'Debe ingresar una observación para solicitar la corrección.')
            return redirect_after_action(request, devolucion)

        ahora = timezone.now()
        estado_anterior = devolucion.get_estado_display()

        devolucion.estado = 'OBSERVADO'
        devolucion.respuesta_cliente = observacion
        devolucion.respondido_por = request.user
        devolucion.fecha_observacion_cliente = ahora

        if not devolucion.fecha_en_revision:
            devolucion.fecha_en_revision = ahora

        devolucion.save(update_fields=[
            'estado',
            'respuesta_cliente',
            'respondido_por',
            'fecha_observacion_cliente',
            'fecha_en_revision',
            'actualizado_en',
        ])

        registrar_historial(
            devolucion=devolucion,
            titulo='Corrección solicitada por cliente',
            descripcion=f'Estado: {estado_anterior} → Pendiente corrección. Observación: {observacion}',
            usuario=request.user,
            tipo='OBSERVADO'
        )

        messages.success(request, 'Solicitud de corrección enviada correctamente.')
        return redirect_after_action(request, devolucion)

    return redirect_after_action(request, devolucion)


@login_required
def return_send_review(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if es_cliente(request.user):
        raise PermissionDenied

    if devolucion.estado != 'OBSERVADO':
        messages.warning(request, 'Solo puedes enviar a revisión una devolución pendiente de corrección.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        ahora = timezone.now()

        devolucion.estado = 'EN_REVISION'
        devolucion.fecha_en_revision = ahora
        devolucion.save(update_fields=[
            'estado',
            'fecha_en_revision',
            'actualizado_en',
        ])

        registrar_historial(
            devolucion=devolucion,
            titulo='Corrección enviada a revisión',
            descripcion='El administrador realizó la corrección y envió la devolución nuevamente a revisión del cliente.',
            usuario=request.user,
            tipo='EN_REVISION'
        )

        messages.success(request, 'Devolución enviada nuevamente a revisión.')
        return redirect_after_action(request, devolucion)

    return redirect_after_action(request, devolucion)


@login_required
def return_close(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    if es_cliente(request.user):
        raise PermissionDenied

    if devolucion.estado != 'APROBADO':
        messages.warning(request, 'Solo se pueden cerrar devoluciones aprobadas.')
        return redirect_after_action(request, devolucion)

    if request.method == 'POST':
        ahora = timezone.now()

        devolucion.estado = 'CERRADO'
        devolucion.fecha_cierre = ahora
        devolucion.cerrado_por = request.user
        devolucion.save(update_fields=[
            'estado',
            'fecha_cierre',
            'cerrado_por',
            'actualizado_en',
        ])

        registrar_historial(
            devolucion=devolucion,
            titulo='Devolución cerrada',
            descripcion='La devolución fue cerrada. Caso ingresado y finalizado en WMS.',
            usuario=request.user,
            tipo='CERRADO'
        )

        messages.success(request, 'Devolución cerrada correctamente.')
        return redirect_after_action(request, devolucion)

    return redirect_after_action(request, devolucion)


# ==========================================================
# USUARIOS
# ==========================================================

@login_required
def user_create(request):
    if not request.user.is_staff and not request.user.is_superuser:
        raise PermissionDenied

    if request.method == 'POST':
        form_usuario = ClienteUserCreateForm(request.POST)

        if form_usuario.is_valid():
            usuario = form_usuario.save()
            messages.success(
                request,
                f'Usuario cliente "{usuario.username}" creado correctamente.'
            )
            return redirect('returns:return_list')

        base_queryset, devoluciones, query, estado, fecha_desde, fecha_hasta = obtener_devoluciones_filtradas(request)

        return render(request, 'returns/return_list.html', {
            'devoluciones': devoluciones,
            'query': query,
            'estado': estado,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado_choices': ReturnRequest.ESTADO_CHOICES,

            'form': ReturnRequestForm(user=request.user),
            'user_form': form_usuario,

            'open_user_modal': True,
            'es_cliente': es_cliente(request.user),

            'total_devoluciones': base_queryset.count(),
            'total_recibidas': devoluciones.filter(estado='RECIBIDO').count(),
            'total_revision': devoluciones.filter(estado='EN_REVISION').count(),
            'total_observadas': devoluciones.filter(estado='OBSERVADO').count(),
            'total_aprobadas': devoluciones.filter(estado='APROBADO').count(),
            'total_rechazadas': devoluciones.filter(estado='RECHAZADO').count(),
            'total_cerradas': devoluciones.filter(estado='CERRADO').count(),
            'total_filtradas': devoluciones.count(),
        })

    return redirect('returns:return_list')


# ==========================================================
# FEEDS AJAX / LIVE
# ==========================================================

@login_required
def return_status_feed(request):
    devoluciones = ReturnRequest.objects.all()

    data = []

    for item in devoluciones:
        data.append({
            'id': item.id,
            'estado': item.estado,
            'estado_display': item.get_estado_display(),
        })

    payload = {
        'items': data,
        'counts': {
            'total': devoluciones.count(),
            'recibidas': devoluciones.filter(estado='RECIBIDO').count(),
            'revision': devoluciones.filter(estado='EN_REVISION').count(),
            'observadas': devoluciones.filter(estado='OBSERVADO').count(),
            'aprobadas': devoluciones.filter(estado='APROBADO').count(),
            'rechazadas': devoluciones.filter(estado='RECHAZADO').count(),
            'cerradas': devoluciones.filter(estado='CERRADO').count(),
        }
    }

    return JsonResponse(payload)


@login_required
def return_live_feed(request):
    base_queryset, devoluciones, query, estado, fecha_desde, fecha_hasta = obtener_devoluciones_filtradas(request)
    cliente = es_cliente(request.user)

    items = []

    for item in devoluciones:
        series = list(item.series.values_list('serie', flat=True))

        items.append({
            'id': item.id,
            'correlativo': item.correlativo,
            'fecha_recepcion': item.fecha_recepcion.strftime('%b %d, %Y'),
            'numero_documento': item.numero_documento or '',
            'documento_ingreso': item.documento_ingreso or 'None',
            'cliente': item.cliente or '',
            'sku': item.sku or '',
            'cantidad': item.cantidad,
            'serie': resumen_series_devolucion(item),
            'series': series,
            'estado_inventario': item.numero_dcto_estado or '-',
            'responsable': item.responsable or '-',
            'estado': item.estado,
            'estado_display': item.get_estado_display(),
            'detail_url': reverse('returns:return_detail', args=[item.pk]),
            'edit_url': reverse('returns:return_update', args=[item.pk]),
            'is_cliente': cliente,
            'respond_url': reverse('returns:return_respond', args=[item.pk]),
            'observe_url': reverse('returns:return_observe', args=[item.pk]),
            'reject_url': reverse('returns:return_reject', args=[item.pk]),
            'close_url': reverse('returns:return_close', args=[item.pk]),
            'send_review_url': reverse('returns:return_send_review', args=[item.pk]),
            'update_url': reverse('returns:return_update', args=[item.pk]),
            'timeline_url': reverse('returns:return_timeline_feed', args=[item.pk]),
            'fecha_iso': item.fecha_recepcion.strftime('%Y-%m-%d'),
            'codigo_numerico': item.codigo_numerico or '',
            'ubicacion': item.ubicacion or '',
            'observaciones': item.observaciones or '',
            'galeria': obtener_galeria_devolucion(item),
        })

    return JsonResponse({
        'items': items,
        'counts': {
            'total': devoluciones.count(),
            'recibidas': devoluciones.filter(estado='RECIBIDO').count(),
            'revision': devoluciones.filter(estado='EN_REVISION').count(),
            'observadas': devoluciones.filter(estado='OBSERVADO').count(),
            'aprobadas': devoluciones.filter(estado='APROBADO').count(),
            'rechazadas': devoluciones.filter(estado='RECHAZADO').count(),
            'cerradas': devoluciones.filter(estado='CERRADO').count(),
            'filtradas': devoluciones.count(),
        }
    })


@login_required
def return_timeline_feed(request, pk):
    devolucion = get_object_or_404(ReturnRequest, pk=pk)

    def format_date(value):
        if not value:
            return 'Pendiente'

        return timezone.localtime(value).strftime('%d/%m/%Y %H:%M')

    def user_label(user):
        if not user:
            return 'Sistema'

        return user.get_full_name() or user.username

    def icon_by_type(tipo):
        icons = {
            'CREACION': 'bi-plus-circle',
            'EN_REVISION': 'bi-eye',
            'OBSERVADO': 'bi-chat-square-text',
            'RESPUESTA_CLIENTE': 'bi-send',
            'CAMBIO_CLIENTE': 'bi-pencil-square',
            'APROBADO': 'bi-check-circle',
            'RECHAZADO': 'bi-x-circle',
            'CERRADO': 'bi-lock',
            'SERIES': 'bi-upc-scan',
            'ADJUNTOS_INTERNOS': 'bi-paperclip',
        }

        return icons.get(tipo, 'bi-clock-history')

    def color_by_type(tipo):
        if tipo in ['APROBADO', 'CERRADO']:
            return 'success'

        if tipo == 'RECHAZADO':
            return 'danger'

        if tipo in ['OBSERVADO', 'EN_REVISION']:
            return 'warning'

        return 'primary'

    items = []

    items.append({
        'tipo': 'CREACION',
        'titulo': 'Creación de devolución',
        'descripcion': f'Devolución creada por {devolucion.responsable or "Sistema"}.',
        'fecha': format_date(devolucion.creado_en),
        'usuario': user_label(devolucion.creado_por) if devolucion.creado_por else devolucion.responsable or 'Sistema',
        'icon': icon_by_type('CREACION'),
        'color': color_by_type('CREACION'),
    })

    for item in devolucion.historial.all().order_by('creado_en', 'id'):
        items.append({
            'tipo': item.tipo,
            'titulo': item.titulo,
            'descripcion': item.descripcion,
            'fecha': format_date(item.creado_en),
            'usuario': user_label(item.creado_por),
            'icon': icon_by_type(item.tipo),
            'color': color_by_type(item.tipo),
        })

    return JsonResponse({
        'id': devolucion.id,
        'correlativo': devolucion.correlativo,
        'items': items,
    })


# ==========================================================
# EXPORTACIÓN EXCEL
# ==========================================================

@login_required
def return_export_xlsx(request):
    base_queryset, devoluciones, query, estado, fecha_desde, fecha_hasta = obtener_devoluciones_filtradas(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Devoluciones"

    headers = [
        "Correlativo",
        "Fecha recepción",
        "N° Documento",
        "Documento ingreso",
        "Cliente",
        "SKU",
        "Cantidad",
        "Series",
        "Estado inventario",
        "Estado",
        "Responsable",
        "Observaciones",
    ]

    ws.append(headers)

    for item in devoluciones:
        ws.append([
            item.correlativo or "",
            item.fecha_recepcion if item.fecha_recepcion else "",
            item.numero_documento or "",
            item.documento_ingreso or "",
            item.cliente or "",
            item.sku or "",
            item.cantidad or 0,
            texto_series_exportacion(item),
            item.numero_dcto_estado or "",
            item.get_estado_display(),
            item.responsable or "",
            item.observaciones or "",
        ])

    header_fill = PatternFill("solid", fgColor="0B4C6F")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "dd/mm/yyyy"

    column_widths = {
        "A": 16,
        "B": 18,
        "C": 28,
        "D": 26,
        "E": 28,
        "F": 18,
        "G": 12,
        "H": 45,
        "I": 22,
        "J": 18,
        "K": 24,
        "L": 40,
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 24

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="devoluciones.xlsx"'

    wb.save(response)
    return response
